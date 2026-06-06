#!/usr/bin/env python3
"""Import 26SM teacher availability from Google Form Excel export."""

from __future__ import annotations

import json
import os
import re
import ssl
import sys
import urllib.error
import urllib.request
from datetime import date, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
EXCEL_DEFAULT = Path.home() / "Downloads/【26年】明學教育導師暑期開班時間表收集 (回應).xlsx"

YEAR_ID = "794b3432-cbef-465b-b2b3-362f7b47f30c"
RANGE_FROM = date(2026, 7, 1)
RANGE_TO = date(2026, 8, 31)

SLOTS = [
    "09:00–10:15",
    "10:15–11:30",
    "11:30–12:45",
    "12:45–14:00",
    "14:00–15:15",
    "15:15–16:30",
    "16:30–17:45",
    "17:45–19:00",
    "19:00–20:15",
    "20:15–21:30",
]

START_TO_INDEX = {
    "09:00": 0,
    "10:15": 1,
    "11:30": 2,
    "12:45": 3,
    "14:00": 4,
    "15:15": 5,
    "16:30": 6,
    "17:45": 7,
    "19:00": 8,
    "20:15": 9,
}

CN_NUM = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
}

WEEKDAY_COLS = {
    4: 0,  # Mon
    5: 1,
    6: 2,
    7: 3,
    8: 4,
    9: 5,
    10: 6,  # Sun
}

SKIP_TEACHERS = {"thom cheong"}


def load_env() -> tuple[str, str]:
    url = key = ""
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        if line.startswith("VITE_SUPABASE_URL="):
            url = line.split("=", 1)[1].strip()
        if line.startswith("VITE_SUPABASE_ANON_KEY="):
            key = line.split("=", 1)[1].strip()
    if not url or not key:
        raise SystemExit("Missing Supabase env in .env")
    return url, key


def _ssl_ctx() -> ssl.SSLContext:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


def api_get(base: str, key: str, path: str) -> list | dict:
    req = urllib.request.Request(
        f"{base}{path}",
        headers={"apikey": key, "Authorization": f"Bearer {key}"},
    )
    with urllib.request.urlopen(req, context=_ssl_ctx()) as resp:
        return json.loads(resp.read().decode())


def api_post(base: str, key: str, path: str, rows: list[dict]) -> None:
    data = json.dumps(rows).encode()
    req = urllib.request.Request(
        f"{base}{path}",
        data=data,
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
    )
    with urllib.request.urlopen(req, context=_ssl_ctx()) as resp:
        resp.read()


def cn_period_to_index(label: str) -> int | None:
    m = re.search(r"第([一二三四五六七八九十]+)節", label)
    if not m:
        return None
    s = m.group(1)
    if s == "十":
        n = 10
    elif len(s) == 2 and s[0] == "十":
        n = 10 + CN_NUM.get(s[1], 0)
    else:
        n = CN_NUM.get(s, 0)
    if 1 <= n <= 10:
        return n - 1
    return None


def parse_slot_indices(cell: object) -> list[int]:
    if cell is None:
        return []
    text = str(cell)
    indices: set[int] = set()
    for part in re.split(r",|，", text):
        idx = cn_period_to_index(part)
        if idx is not None:
            indices.add(idx)
            continue
        tm = re.search(r"(\d{1,2}:\d{2})\s*[-–—~]\s*(\d{1,2}:\d{2})", part)
        if tm:
            start = tm.group(1).zfill(5) if len(tm.group(1)) == 4 else tm.group(1)
            if len(start) == 4:
                start = "0" + start
            if start in START_TO_INDEX:
                indices.add(START_TO_INDEX[start])
    return sorted(indices)


def daterange(a: date, b: date) -> set[str]:
    out: set[str] = set()
    cur = a
    while cur <= b:
        out.add(cur.isoformat())
        cur += timedelta(days=1)
    return out


def parse_md(m: int, d: int, default_year: int = 2026) -> date:
    year = default_year
    if m == 9 and d <= 3:
        year = 2026
    return date(year, m, d)


def parse_exceptions(text: object) -> set[str]:
    if not text:
        return set()
    s = str(text).strip().lower()
    if s in {"沒有", "未有", "暫無", "暫時沒有", "no", "none", "n/a"}:
        return set()
    if "畢業" in s:
        return set()

    excluded: set[str] = set()

    for m in re.finditer(
        r"(\d{1,2})\s*月\s*(\d{1,2})\s*[、,，]\s*(\d{1,2})",
        str(text),
    ):
        month = int(m.group(1))
        excluded |= {parse_md(month, int(m.group(2))).isoformat(), parse_md(month, int(m.group(3))).isoformat()}

    for m in re.finditer(
        r"(\d{1,2})\s*月\s*(\d{1,2})",
        str(text),
    ):
        if re.search(rf"{m.group(1)}\s*月\s*{m.group(2)}\s*[、,，]", str(text)):
            continue
        month = int(m.group(1))
        excluded.add(parse_md(month, int(m.group(2))).isoformat())

    for m in re.finditer(
        r"(\d{1,2})\s*/\s*(\d{1,2})\s*[-–—~至到]+\s*(\d{1,2})\s*/\s*(\d{1,2})",
        str(text),
    ):
        a = parse_md(int(m.group(2)), int(m.group(1)))
        b = parse_md(int(m.group(4)), int(m.group(3)))
        if b < a and int(m.group(4)) < int(m.group(2)):
            b = parse_md(int(m.group(4)), int(m.group(3)), 2026)
        if b < a:
            b = parse_md(int(m.group(4)), int(m.group(3)))
        excluded |= daterange(min(a, b), max(a, b))

    return {d for d in excluded if RANGE_FROM.isoformat() <= d <= RANGE_TO.isoformat()}


def dates_for_weekday(weekday: int) -> list[str]:
    # weekday: 0=Mon .. 6=Sun; Python date.weekday() same
    out: list[str] = []
    cur = RANGE_FROM
    while cur <= RANGE_TO:
        if cur.weekday() == weekday:
            out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out


def match_teacher(name: str, teachers: list[dict]) -> dict | None:
    n = re.sub(r"\s+", " ", name.strip().lower())
    if n in SKIP_TEACHERS:
        return None
    for t in teachers:
        full = t["full_name"].strip().lower()
        eng = (t.get("english_name") or "").strip().lower()
        if n == full or (eng and n == eng):
            return t
    for t in teachers:
        if t["full_name"].strip().lower().replace(" ", "") == n.replace(" ", ""):
            return t
    return None


def main() -> None:
    dry_run = "--dry-run" in sys.argv
    excel_path = EXCEL_DEFAULT
    for arg in sys.argv[1:]:
        if not arg.startswith("-"):
            excel_path = Path(arg)

    base, key = load_env()
    teachers = api_get(base, key, "/rest/v1/teachers?select=id,full_name,english_name")
    existing = api_get(
        base,
        key,
        f"/rest/v1/teacher_availability_slots?select=teacher_id,available_date,time_slot"
        f"&academic_year_id=eq.{YEAR_ID}"
        f"&available_date=gte.{RANGE_FROM.isoformat()}"
        f"&available_date=lte.{RANGE_TO.isoformat()}",
    )
    existing_keys = {
        (r["teacher_id"], r["available_date"][:10], r["time_slot"].replace("-", "–"))
        for r in existing
    }

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["表格回應 1"]

    to_insert: list[dict] = []
    report: list[str] = []
    unmatched: list[str] = []

    for row in range(2, ws.max_row + 1):
        raw_name = ws.cell(row, 2).value
        if not raw_name:
            continue
        teacher = match_teacher(str(raw_name), teachers)
        if not teacher:
            unmatched.append(str(raw_name).strip())
            continue

        excluded = parse_exceptions(ws.cell(row, 11).value)

        teacher_slots = 0
        for col, weekday in WEEKDAY_COLS.items():
            indices = parse_slot_indices(ws.cell(row, col).value)
            if not indices:
                continue
            for ymd in dates_for_weekday(weekday):
                if ymd in excluded:
                    continue
                for idx in indices:
                    slot = SLOTS[idx]
                    key_tuple = (teacher["id"], ymd, slot)
                    if key_tuple in existing_keys:
                        continue
                    existing_keys.add(key_tuple)
                    row_data = {
                        "teacher_id": teacher["id"],
                        "academic_year_id": YEAR_ID,
                        "available_date": ymd,
                        "time_slot": slot,
                        "status": "可分配",
                        "notes": "26SM 問卷匯入",
                    }
                    to_insert.append(row_data)
                    teacher_slots += 1

        report.append(f"{teacher['full_name']}: +{teacher_slots} slots, excluded {len(excluded)} dates")

    print(f"Excel: {excel_path}")
    print(f"Prepared {len(to_insert)} new availability slots")
    for line in report:
        print(" ", line)
    if unmatched:
        print("Unmatched names:", ", ".join(unmatched))

    if dry_run:
        print("\nDry run — no writes.")
        return

    batch = 200
    for i in range(0, len(to_insert), batch):
        chunk = to_insert[i : i + batch]
        try:
            api_post(base, key, "/rest/v1/teacher_availability_slots", chunk)
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            print(f"Insert failed at batch {i}: {e.code} {body}", file=sys.stderr)
            raise SystemExit(1)
        print(f"Inserted {min(i + batch, len(to_insert))}/{len(to_insert)}")

    print("Done.")


if __name__ == "__main__":
    main()
