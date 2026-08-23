#!/usr/bin/env python3
"""匯出 2526（Notion 六月名單）＋ 26SM 系統報讀：學生 × 科目 × 老師，供前線覆核。"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = ROOT / "docs/generated/2627/2526_26SM_SUBJECT_ENROLLMENTS_REVIEW.xlsx"

NOTION_TEACHER_MARKERS: list[tuple[str, str]] = [
    ("Cheryl Ng", "Cheryl Ng"),
    ("Rafael", "Rafael Ling"),
    ("CFAN", "Christine Fan"),
    ("CYNG", "Cyndi Ng"),
    ("JCHU", "Judy Chu"),
    ("PHBE", "Phoebe Tam"),
    ("PHEB", "Phoebe Tam"),
    ("THOM", "Thom Cheong"),
    ("LIAM", "Liam Lai"),
    ("NKWO", "Natalie Kwok"),
    ("SHEK", "Billy Shek"),
    ("TIMC", "Tim Cheung"),
    ("RALI", "Rafael Ling"),
    ("LING", "Rafael Ling"),
    ("JLAU", "Jackson Lau"),
    ("MYU", "Mark Yu"),
    ("Rain", "Rain"),
]


SQL_LIVE = ROOT / "scripts/sql/export_2526_26sm_live.sql"
SQL_LEGACY = ROOT / "scripts/sql/export_2526_legacy.sql"


def query_sql(sql_path: Path) -> list[dict[str, Any]]:
    supabase = shutil.which("supabase")
    if not supabase:
        local = Path.home() / ".local/bin/supabase"
        supabase = str(local) if local.is_file() else "supabase"
    env = {**os.environ, "PATH": os.environ.get("PATH", "") + os.pathsep + str(Path.home() / ".local/bin")}
    completed = subprocess.run(
        [supabase, "db", "query", "--linked", "-f", str(sql_path), "--output-format", "json"],
        cwd=ROOT,
        env=env,
        capture_output=True,
        text=True,
        check=False,
    )
    if completed.returncode != 0:
        raise SystemExit(completed.stderr or completed.stdout or "supabase db query failed")
    raw = completed.stdout.strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"JSON parse failed for {sql_path.name}: {exc}\n{raw[:400]}") from exc
    if isinstance(data, dict) and isinstance(data.get("rows"), list):
        return data["rows"]
    if isinstance(data, list):
        return data
    raise SystemExit(f"Unexpected query output from {sql_path.name}: {type(data)} keys={list(data)[:8] if isinstance(data, dict) else ''}")


def text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def teacher_from_notion_label(label: str) -> str:
    compact = re.sub(r"\s+", "", label)
    for marker, name in NOTION_TEACHER_MARKERS:
        if compact.upper().endswith(re.sub(r"\s+", "", marker).upper()) or marker in label:
            return name
    return ""


def sort_key(row: dict[str, str]) -> tuple[str, str, str, str]:
    code = row["學號"] or "zzz"
    year = "0" if row["學年"] == "2526" else "1"
    return (code, year, row["科目"], row["老師"])


def add_sheet_note(ws: Worksheet) -> None:
    ws["A1"] = "2526／26SM 報讀科目覆核表"
    ws["A1"].font = Font(name="PingFang TC", size=16, bold=True, color="1A2540")
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"產出日：{date.today().isoformat()}。"
        "2526＝Notion 匯入（約 2026-06 仍在讀的科目，唔係全年報讀）。"
        "26SM＝系統就讀中／已退讀。"
        "請前線核對：有冇學生 5 月或之前讀過 2526，但下面沒有 2526 列——有則補在「請補漏」工作表。"
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 48
    ws["A2"].font = Font(name="PingFang TC", size=11, color="57534E")


def style_header(ws: Worksheet, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="243357")
    font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")


def write_table(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[str]],
    start_row: int,
    table_name: str,
) -> None:
    thin = Border(
        left=Side(style="thin", color="E5E0D5"),
        right=Side(style="thin", color="E5E0D5"),
        top=Side(style="thin", color="E5E0D5"),
        bottom=Side(style="thin", color="E5E0D5"),
    )
    for col, title in enumerate(headers, 1):
        ws.cell(start_row, col, title)
    style_header(ws, start_row, len(headers))
    body_font = Font(name="PingFang TC", size=11, color="1C2438")
    for r_i, values in enumerate(rows, start_row + 1):
        for c_i, value in enumerate(values, 1):
            cell = ws.cell(r_i, c_i, value)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
        if (r_i - start_row) % 2 == 0:
            fill = PatternFill("solid", fgColor="F7F5F0")
            for c_i in range(1, len(headers) + 1):
                ws.cell(r_i, c_i).fill = fill
    last_row = max(start_row + len(rows), start_row + 1)
    last_col = get_column_letter(len(headers))
    table = Table(
        displayName=table_name,
        ref=f"A{start_row}:{last_col}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{start_row + 1}"
    widths = [14, 12, 10, 22, 16, 28]
    for i, width in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_rows() -> list[dict[str, str]]:
    live = query_sql(SQL_LIVE)
    legacy = query_sql(SQL_LEGACY)
    out: list[dict[str, str]] = []

    for row in live:
        year = text(row.get("academic_year"))
        status = text(row.get("status")) or "就讀中"
        source = "26SM 系統報讀" if year == "26SM" else "2526 系統（私人課程殘留）"
        if year == "26SM" and status == "已退讀":
            source = "26SM 系統報讀（已退讀）"
        out.append(
            {
                "姓名": text(row.get("student_name")),
                "學號": text(row.get("student_code")),
                "學年": year,
                "科目": text(row.get("subject")),
                "老師": text(row.get("teacher_name")),
                "資料來源": source,
            }
        )

    for row in legacy:
        label = text(row.get("source_subject_label"))
        out.append(
            {
                "姓名": text(row.get("student_name")),
                "學號": text(row.get("student_code")),
                "學年": "2526",
                "科目": text(row.get("subject")),
                "老師": teacher_from_notion_label(label),
                "資料來源": "2526 Notion（6月仍在讀）",
            }
        )

    out.sort(key=sort_key)
    return out


def main() -> int:
    rows = build_rows()
    headers = ["姓名", "學號", "學年", "科目", "老師", "資料來源"]
    values = [[r[h] for h in headers] for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "報讀清單"
    add_sheet_note(ws)
    write_table(ws, headers, values, start_row=4, table_name="EnrollmentReview")
    ws.row_dimensions[4].height = 22
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "4:4"
    ws.oddHeader.left.text = "明學教育　2526／26SM 報讀科目覆核"

    blank = wb.create_sheet("請補漏")
    blank["A1"] = "請補：2526 年 5 月或之前有讀、但「報讀清單」沒有 2526 列的學生"
    blank["A1"].font = Font(name="PingFang TC", size=14, bold=True, color="1A2540")
    blank.merge_cells("A1:E1")
    blank["A2"] = (
        "一行一科。學年請填 2526。補完後交回，方便分開「2526 完全沒讀／曾經讀過」兩種 2627 宣傳方法。"
    )
    blank["A2"].font = Font(name="PingFang TC", size=11, color="57534E")
    blank.merge_cells("A2:E2")
    write_table(
        blank,
        ["姓名", "學號", "學年", "科目", "老師"],
        [["", "", "2526", "", ""] for _ in range(30)],
        start_row=4,
        table_name="MissingEnrollmentReview",
    )

    note = wb.create_sheet("說明")
    note["A1"] = "口徑"
    note["A1"].font = Font(name="PingFang TC", size=16, bold=True)
    lines = [
        "1. 2526 學年未用本系統；7 月啟用後，只匯入「學生 × 曾讀科目」，不是全年正式報讀。",
        "2. Notion 名單只反映約 2026-06 仍在讀的人。考試結束後 6 月停讀、但 5 月或之前有讀的人，可能不在表內。",
        "3. 26SM 列來自系統報讀（就讀中及已退讀）。",
        "4. 「2526 系統（私人課程殘留）」是現時系統仍掛 2526 的少數私人課程，不是小組班全年紀錄。",
        "5. 2526 老師多由 Notion 班名尾碼還原（如 MYU＝Mark Yu）；班名無老師碼則老師欄空白，請同事補。",
        "6. 排序：學號 → 2526 先於 26SM → 科目。",
        f"7. 本檔列數：{len(rows)}。",
    ]
    for i, line in enumerate(lines, 3):
        note[f"A{i}"] = line
        note[f"A{i}"].font = Font(name="PingFang TC", size=12)
        note[f"A{i}"].alignment = Alignment(wrap_text=True)
        note.row_dimensions[i].height = 28
    note.column_dimensions["A"].width = 110

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    n_2526 = sum(1 for r in rows if r["學年"] == "2526")
    n_26sm = sum(1 for r in rows if r["學年"] == "26SM")
    students = {r["學號"] or r["姓名"] for r in rows}
    print(
        json.dumps(
            {
                "path": str(OUT_PATH),
                "rows": len(rows),
                "students": len(students),
                "rows_2526": n_2526,
                "rows_26sm": n_26sm,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
