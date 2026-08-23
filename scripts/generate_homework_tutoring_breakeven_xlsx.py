#!/usr/bin/env python3
"""Generate homework-tutoring break-even workbook with live Excel formulas."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = ROOT / "docs/generated/2627/HOMEWORK_TUTORING_BREAKEVEN.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF3CD")
HEADER = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
LABEL_FONT = Font(name="Calibri", size=11)
INPUT_FONT = Font(name="Calibri", size=11, color="7A5C00")
FORMULA_FILL = PatternFill("solid", fgColor="F4F6F8")
TOTAL_FILL = PatternFill("solid", fgColor="D6EAF8")
OK_FILL = PatternFill("solid", fgColor="C8E6C9")
BAD_FILL = PatternFill("solid", fgColor="FFCDD2")
THIN = Border(
    left=Side(style="thin", color="C5CDD6"),
    right=Side(style="thin", color="C5CDD6"),
    top=Side(style="thin", color="C5CDD6"),
    bottom=Side(style="thin", color="C5CDD6"),
)
WRAP = Alignment(wrap_text=True, vertical="center")

MONTHS = [
    ("9 月", 22, 1),
    ("10 月", 20, 1),
    ("11 月", 21, 1),
    ("12 月", 15, 0.75),
    ("1 月", 20, 1),
    ("2 月", 15, 0.75),
    ("3 月", 20, 1),
    ("4 月", 21, 1),
    ("5 月", 20, 1),
    ("6 月", 21, 1),
]


def style_input(cell, number_format: str | None = None) -> None:
    cell.fill = YELLOW
    cell.font = INPUT_FONT
    cell.border = THIN
    cell.alignment = Alignment(vertical="center")
    if number_format:
        cell.number_format = number_format


def style_formula(cell, number_format: str | None = None) -> None:
    cell.fill = FORMULA_FILL
    cell.font = LABEL_FONT
    cell.border = THIN
    if number_format:
        cell.number_format = number_format


def style_header_row(ws: Worksheet, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def build_params(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "參數"
    ws.sheet_properties.tabColor = "1F4E79"
    ws["A1"] = "功輔回本假設（黃格可改；灰格是公式，勿覆寫）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A2"] = (
        "場地按 23/31 歸功輔、8/31 歸專科週末。Katie 全職月薪 × 佔比計入功輔。"
        "小學／中學分頁會用本頁數字自動重算。"
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 36

    ws["A4"] = "場地（17D＋17E）"
    ws["A4"].font = SECTION_FONT
    rows = [
        (5, "月租（港元）", 15000, "#,##0"),
        (6, "管理費（港元／月）", 4881, "#,##0"),
        (7, "差餉（港元／季）", 3660, "#,##0"),
    ]
    for r, label, val, fmt in rows:
        ws.cell(r, 1, label).font = LABEL_FONT
        ws.cell(r, 2, val)
        style_input(ws.cell(r, 2), fmt)

    ws["A8"] = "差餉攤每月"
    ws["B8"] = "=B7/3"
    style_formula(ws["B8"], "#,##0.00")
    ws["A9"] = "場地每月合計"
    ws["B9"] = "=B5+B6+B8"
    style_formula(ws["B9"], "#,##0.00")

    ws["A10"] = "功輔佔日數"
    ws["B10"] = 23
    style_input(ws["B10"], "0")
    ws["A11"] = "一個月日數"
    ws["B11"] = 31
    style_input(ws["B11"], "0")
    ws["A12"] = "專科週末佔日數"
    ws["B12"] = "=B11-B10"
    style_formula(ws["B12"], "0")
    ws["A13"] = "功輔場地（佔比分攤）"
    ws["B13"] = "=B9*B10/B11"
    style_formula(ws["B13"], "#,##0.00")
    ws["A14"] = "專科週末場地"
    ws["B14"] = "=B9*B12/B11"
    style_formula(ws["B14"], "#,##0.00")
    ws["C13"] = "← 兩室功輔合計"
    ws["C14"] = "← 不入功輔回本"

    ws["A16"] = "人事"
    ws["A16"].font = SECTION_FONT
    ws["A17"] = "Katie 全職月薪"
    ws["B17"] = 20000
    style_input(ws["B17"], "#,##0")
    ws["A18"] = "功輔佔 Katie"
    ws["B18"] = 0.3
    style_input(ws["B18"], "0%")
    ws["A19"] = "Katie 計入功輔"
    ws["B19"] = "=B17*B18"
    style_formula(ws["B19"], "#,##0.00")
    ws["A20"] = "資深 PT 日薪"
    ws["B20"] = 350
    style_input(ws["B20"], "#,##0")
    ws["A21"] = "普通 PT 日薪"
    ws["B21"] = 210
    style_input(ws["B21"], "#,##0")

    ws["A23"] = "兩室分攤場地及 Katie（合計應為 100%）"
    ws["A23"].font = SECTION_FONT
    ws.merge_cells("A23:C23")
    ws["A24"] = "小學分攤"
    ws["B24"] = 0.5
    style_input(ws["B24"], "0%")
    ws["A25"] = "中學分攤"
    ws["B25"] = 0.5
    style_input(ws["B25"], "0%")
    ws["A26"] = "分攤合計（應＝100%）"
    ws["B26"] = "=B24+B25"
    style_formula(ws["B26"], "0%")
    ws["A27"] = "典型月平日（掃描表用）"
    ws["B27"] = 20
    style_input(ws["B27"], "0")

    ws["A29"] = "校曆（可改平日數或學費比例）"
    ws["A29"].font = SECTION_FONT
    ws.merge_cells("A29:C29")
    ws["A30"] = "月份"
    ws["B30"] = "平日開放"
    ws["C30"] = "學費比例"
    style_header_row(ws, 30, 3)
    for i, (name, days, share) in enumerate(MONTHS):
        r = 31 + i
        ws.cell(r, 1, name).border = THIN
        ws.cell(r, 2, days)
        style_input(ws.cell(r, 2), "0")
        ws.cell(r, 3, share)
        style_input(ws.cell(r, 3), "0%")

    ws["A42"] = "學年平日合計"
    ws["B42"] = "=SUM(B31:B40)"
    style_formula(ws["B42"], "0")
    ws["A43"] = "學費月數合計（12／2 月 75% 會計 0.75）"
    ws["B43"] = "=SUM(C31:C40)"
    style_formula(ws["B43"], "0.00")

    ws["A45"] = "編制規則"
    ws["A45"].font = SECTION_FONT
    ws.merge_cells("A45:C45")
    ws["A46"] = (
        "中學：人數>0 → ROUNDUP(人數/7,0) 位資深（1–7 一位，8–14 兩位）。\n"
        "小學：人數>0 → 1 位資深，帶滿「資深每位應對」後加普通；"
        "普通人數＝ROUNDUP((人數−資深每位)/普通每位,0)。容量在小學頁黃格改。"
    )
    ws["A46"].alignment = WRAP
    ws.merge_cells("A46:C48")
    ws.row_dimensions[46].height = 28
    ws.row_dimensions[47].height = 28
    ws.row_dimensions[48].height = 28

    ws["A50"] = "小學資深每位應對（來自小學頁）"
    ws["B50"] = "=小學!B7"
    style_formula(ws["B50"], "0")
    ws["A51"] = "小學普通每位應對（來自小學頁）"
    ws["B51"] = "=小學!D7"
    style_formula(ws["B51"], "0")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A31"
    ws.row_dimensions[1].height = 24


def add_gap_cf(ws: Worksheet, gap_col: str, start: int, end: int) -> None:
    rng = f"{gap_col}{start}:{gap_col}{end}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=OK_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0"], fill=BAD_FILL)
    )


def build_room_sheet(
    wb: openpyxl.Workbook,
    title: str,
    *,
    default_n: int,
    default_fee: int,
    share_cell: str,
    is_primary: bool,
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = "2E7D32" if is_primary else "1565C0"

    ws["A1"] = f"{title}功輔室回本"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    rule = (
        "小學編制：人數>0 → 1 位資深；超出「資深每位應對」後按「普通每位應對」加普通。"
        if is_primary
        else "中學編制：人數>0 → ROUNDUP(人數/7,0) 位資深（1–7 一位，8–14 兩位）。"
    )
    ws["A2"] = (
        f"黃格可改。場地同 Katie 按「參數」頁分攤。{rule} "
        "每月場地／Katie 不跟該月平日數加減；PT 跟該月平日 × 日薪。"
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 40

    ws["A4"] = "輸入"
    ws["A4"].font = SECTION_FONT
    ws["A5"] = "人數"
    ws["B5"] = default_n
    style_input(ws["B5"], "0")
    ws["A6"] = "月費（全費月份）"
    ws["B6"] = default_fee
    style_input(ws["B6"], "#,##0")
    ws["C6"] = "12 月／2 月自動 × 參數頁學費比例"

    if is_primary:
        ws["A7"] = "資深每位應對學生"
        ws["B7"] = 4
        style_input(ws["B7"], "0")
        ws["C7"] = "普通每位應對學生"
        ws["D7"] = 3
        style_input(ws["D7"], "0")
        ws["E7"] = "人數>0 先 1 位資深；帶滿 B7 後按 D7 加普通。改呢兩格會重算編制／損益／掃描表。"
        ws["E7"].alignment = WRAP
        ws.merge_cells("E7:J7")
        ws.row_dimensions[7].height = 28
        dv = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1="0",
            allow_blank=False,
            errorTitle="請輸入正整數",
            error="資深／普通每位應對學生須為 1 或以上。",
            promptTitle="每位應對",
            prompt="一位老師最多帶幾多學生（正整數）。",
        )
        dv.add("B7")
        dv.add("D7")
        ws.add_data_validation(dv)

    ws["A8"] = "編制（公式）"
    ws["A8"].font = SECTION_FONT
    if is_primary:
        ws["A9"] = "資深老師"
        ws["B9"] = '=IF(B5<=0,0,1)'
        ws["A10"] = "普通老師"
        ws["B10"] = '=IF(OR(B5<=0,$D$7<=0),0,IF(B5<=$B$7,0,ROUNDUP((B5-$B$7)/$D$7,0)))'
    else:
        ws["A9"] = "資深老師"
        ws["B9"] = '=IF(B5<=0,0,ROUNDUP(B5/7,0))'
        ws["A10"] = "普通老師"
        ws["B10"] = 0
    style_formula(ws["B9"], "0")
    style_formula(ws["B10"], "0")
    ws["A11"] = "PT 每日成本"
    ws["B11"] = "=B9*參數!$B$20+B10*參數!$B$21"
    style_formula(ws["B11"], "#,##0.00")
    ws["A12"] = "分攤場地／月"
    ws["B12"] = f"=參數!$B$13*參數!{share_cell}"
    style_formula(ws["B12"], "#,##0.00")
    ws["A13"] = "分攤 Katie／月"
    ws["B13"] = f"=參數!$B$19*參數!{share_cell}"
    style_formula(ws["B13"], "#,##0.00")

    ws["A15"] = "每月損益"
    ws["A15"].font = SECTION_FONT
    headers = [
        "月份",
        "平日",
        "學費比例",
        "場地",
        "Katie",
        "PT",
        "成本",
        "收入",
        "盈虧",
        "累計盈虧",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(16, i, h)
    style_header_row(ws, 16, 10)

    for i in range(10):
        r = 17 + i
        src = 31 + i
        ws.cell(r, 1, f"=參數!A{src}")
        ws.cell(r, 2, f"=參數!B{src}")
        ws.cell(r, 3, f"=參數!C{src}")
        ws.cell(r, 4, "=$B$12")
        ws.cell(r, 5, "=$B$13")
        ws.cell(r, 6, "=$B$11*B{0}".format(r))
        ws.cell(r, 7, f"=D{r}+E{r}+F{r}")
        ws.cell(r, 8, f"=$B$5*$B$6*C{r}")
        ws.cell(r, 9, f"=H{r}-G{r}")
        ws.cell(r, 10, f"=I{r}" if i == 0 else f"=J{r-1}+I{r}")
        for c in range(1, 11):
            style_formula(ws.cell(r, c), None)
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 3).number_format = "0%"
        for c in range(4, 11):
            ws.cell(r, c).number_format = "#,##0"

    ws["A27"] = "學年合計"
    ws["A27"].font = Font(name="Calibri", bold=True, size=11)
    ws["B27"] = "=SUM(B17:B26)"
    ws["C27"] = "=SUM(C17:C26)"
    ws["D27"] = "=SUM(D17:D26)"
    ws["E27"] = "=SUM(E17:E26)"
    ws["F27"] = "=SUM(F17:F26)"
    ws["G27"] = "=SUM(G17:G26)"
    ws["H27"] = "=SUM(H17:H26)"
    ws["I27"] = "=SUM(I17:I26)"
    ws["J27"] = "=I27"
    for c in range(1, 11):
        ws.cell(27, c).fill = TOTAL_FILL
        ws.cell(27, c).font = Font(name="Calibri", bold=True)
        ws.cell(27, c).border = THIN
    ws["B27"].number_format = "0"
    ws["C27"].number_format = "0.00"
    for c in range(4, 11):
        ws.cell(27, c).number_format = "#,##0"

    add_gap_cf(ws, "I", 17, 27)

    ws["A29"] = "人數掃描（典型月：用參數頁「典型月平日」，學費全費）"
    ws["A29"].font = SECTION_FONT
    ws.merge_cells("A29:H29")
    scan_headers = ["人數", "資深", "普通", "PT每日", "典型月成本", "收入", "盈虧", "回本？"]
    for i, h in enumerate(scan_headers, 1):
        ws.cell(30, i, h)
    style_header_row(ws, 30, 8)

    for n in range(0, 21):
        r = 31 + n
        ws.cell(r, 1, n)
        style_formula(ws.cell(r, 1), "0")
        if is_primary:
            ws.cell(r, 2, f"=IF(A{r}<=0,0,1)")
            ws.cell(r, 3, f"=IF(OR(A{r}<=0,$D$7<=0),0,IF(A{r}<=$B$7,0,ROUNDUP((A{r}-$B$7)/$D$7,0)))")
        else:
            ws.cell(r, 2, f"=IF(A{r}<=0,0,ROUNDUP(A{r}/7,0))")
            ws.cell(r, 3, 0)
        ws.cell(r, 4, f"=B{r}*參數!$B$20+C{r}*參數!$B$21")
        ws.cell(r, 5, f"=$B$12+$B$13+D{r}*參數!$B$27")
        ws.cell(r, 6, f"=A{r}*$B$6")
        ws.cell(r, 7, f"=F{r}-E{r}")
        ws.cell(r, 8, f'=IF(A{r}=0,"—",IF(G{r}>=0,"回本","未回本"))')
        for c in range(2, 8):
            style_formula(ws.cell(r, c), None)
        style_formula(ws.cell(r, 8))
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 3).number_format = "0"
        for c in (4, 5, 6, 7):
            ws.cell(r, c).number_format = "#,##0"

    add_gap_cf(ws, "G", 31, 51)
    ws.conditional_formatting.add(
        "A31:H51",
        FormulaRule(formula=["$A31=$B$5"], fill=PatternFill("solid", fgColor="BBDEFB")),
    )

    ws["A53"] = "現設定學年每人學費（已計 12／2 月折扣）"
    ws["B53"] = "=B6*參數!$B$43"
    style_formula(ws["B53"], "#,##0.00")
    ws["A54"] = "現設定學年收入"
    ws["B54"] = "=H27"
    style_formula(ws["B54"], "#,##0")
    ws["A55"] = "現設定學年成本"
    ws["B55"] = "=G27"
    style_formula(ws["B55"], "#,##0")

    widths = {"A": 28, "B": 16, "C": 20, "D": 14, "E": 12, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A17"
    ws.row_dimensions[1].height = 24
    ws.print_title_rows = "1:16"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def main() -> None:
    wb = openpyxl.Workbook()
    build_params(wb)
    build_room_sheet(wb, "小學", default_n=7, default_fee=2800, share_cell="$B$24", is_primary=True)
    build_room_sheet(wb, "中學", default_n=7, default_fee=3200, share_cell="$B$25", is_primary=False)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
